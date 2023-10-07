from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import numbers
from percentileFinder import findValue
from daysNumber import daysNumber


# Ensure the filename, sequence, and cid.
def createReport(cid, startPicture, endPicture, daysNumber):
    endOfMonth = True
    # The error will be logged on an array
    errorLog = []
    fileName = input("Insert your filename: ") + ".xlsx"
    # Example for filename
    # fileName = 'excel_filename.xlsx'

    # Load the workbook
    wb = load_workbook(fileName)

    print('This will be working on row', startPicture, 'until', endPicture)
    cont = input(
        'Please ensure your CID is {} , and please check the filename. Are those okay? y/n '.format(cid))
    if cont == 'n':
        print('Please change on the your cid and/or file.')
        exit()

    #### Sizing stuff ####
    ws = wb.active
    p2e = pixels_to_EMU
    size = XDRPositiveSize2D(p2e(307.2), p2e(86.4))
    size2 = XDRPositiveSize2D(p2e(310.08), p2e(25.92))
    c2e = cm_to_EMU

    # Calculated number of cells width or height from cm into EMUs
    def cellh(x): return c2e((x * 49.77)/99)
    def cellw(x): return c2e((x * (18.65-1.71))/10)

    # Want to place image in row 5 (6 in excel), column 2 (C in excel)
    # Also offset by half a column.
    column = 10
    coloffset = cellw(0.3)
    row = 36
    rowoffset = cellh(0.7)

    # Start The Sequence
    for i in range(startPicture, endPicture):
        # This is for graph sequence
        if daysNumber == 30:
            if i == 31:
                i = 32
        # The cid (customer ID) is started from zero
        if (cid == 0):
            newNameFile = '{} - your_1st_cid.png'.format(i)
        if (cid == 1):
            newNameFile = '{} - your_2nd_cid.png'.format(i)
        if (cid == 2):
            newNameFile = '{} - your_3rd_cid.png'.format(i)
        if (cid == 3):
            newNameFile = '{} - your_4th_cid.png'.format(i)
        if (cid == 4):
            newNameFile = '{} - your_5th_cid.png'.format(i)
        if (cid == 5):
            newNameFile = '{} - your_6th_cid.png'.format(i)
        if (cid == 6):
            newNameFile = '{} - your_7th_cid.png'.format(i)

        #### image location ####
        # The graphic will location is located below:
        img = Image('./imgs/graph/{}'.format(newNameFile))
        # This is for table location
        imgTable = Image('./imgs/table/{}'.format(newNameFile))


        # The value is extracted by OCR on the table, this function will return array
        result = findValue('./imgs/table/{}'.format(newNameFile))
        if newNameFile:
            result = findValue('./imgs/table/{}'.format(newNameFile))
            print(result)

            try:
                result1 = result[0][0].replace(',','')
                result2 = result[0][1].replace(',','')
            except:
                result1 = result[0][0]
                result2 = result[0][1]
                print('unable to fix the value')
            cellRow = 36 + i
            print(result1, result2, 'cell row', cellRow)
            try:
                if (type(result1) == type('string')):
                    ws['D{}'.format(cellRow)] = result1
                    ws['F{}'.format(cellRow)] = result2
                else:
                    ws['D{}'.format(cellRow)] = int(result1)/1000
                    ws['F{}'.format(cellRow)]  = int(result2)/1000  
                    ws['D{}'.format(cellRow)].number_format = '#,##'
                    ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
            except:
                    ws['D{}'.format(cellRow)] = int(0)
                    ws['F{}'.format(cellRow)]  = int(0)  
                    ws['D{}'.format(cellRow)].number_format = '#,##'
                    ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
        

        h, w = img.height, img.width

        if i != 1:
            row = row + 1

        marker1 = AnchorMarker(col=column, colOff=coloffset,
                            row=row, rowOff=rowoffset)

        marker2 = AnchorMarker(col=column, colOff=coloffset,
                            row=row, rowOff=rowoffset + cellh(5.5))

        img.anchor = OneCellAnchor(_from=marker1, ext=size)
        imgTable.anchor = OneCellAnchor(_from=marker2, ext=size2)
        ws.add_image(img)
        ws.add_image(imgTable)
        print('working for {} table'.format(i))

    if endOfMonth == True:
        try:
            if newNameFile:
                newNameFile = newNameFile[2:]
                result = findValue('./imgs/table/32{}'.format(newNameFile))
                # Average field
                result1 = result[1][0]
                result2 = result[1][1]
                # The template of the excel report is required to fill the data on row 69 & 71
                cellRow = 69
                print(result1, result2, 'cell row', cellRow)
                # Divided by 1000 to convert the value
                ws['D{}'.format(cellRow)] = int(result1)/1000
                ws['F{}'.format(cellRow)]  = int(result2)/1000
                ws['D{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
                ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
                
                result1 = result[0][0]
                result2 = result[0][1]
                cellRow = 71
                print(result1, result2, 'cell row', cellRow)
                ws['D{}'.format(cellRow)] = int(result1)/1000
                ws['F{}'.format(cellRow)]  = int(result2)/1000
                ws['D{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER
                ws['F{}'.format(cellRow)].number_format = numbers.FORMAT_NUMBER

        # If the data is not available, it will print 'data not found'
        except:
            print('data not found')

        # 32 series is inteded for the graph start from the first until the end of the month
        img = Image('./imgs/graph/32{}'.format(newNameFile))

        # This is for table image configuration
        imgTable = Image('./imgs/table/32{}'.format(newNameFile))
        marker1 = AnchorMarker(col=10, colOff=coloffset, row=70, rowOff=rowoffset)
        marker2 = AnchorMarker(col=10, colOff=coloffset, row=70, rowOff=rowoffset + cellh(5.5))
        img.anchor = OneCellAnchor(_from=marker1, ext=size)
        imgTable.anchor = OneCellAnchor(_from=marker2, ext=size2)

        # Put the images
        ws.add_image(img)
        ws.add_image(imgTable)
    wb.save('[completed]{}.xlsx'.format(fileName))
    print('please check the log below: {}'.format(errorLog))


# Run the program
createReport(0, 1, 32, 32)

